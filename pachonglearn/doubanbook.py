import os.path
import urllib.parse
import random

import openpyxl
import requests
from bs4 import BeautifulSoup
from urllib import parse
from openpyxl import Workbook

has=[
{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}
]

test_url= "https://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0"
def book_spider(book_tag):
    book_list=[]
    test_url = "https://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0"
    #默认爬取第一页
    tab_url = "https://www.douban.com/tag/"+urllib.parse.quote(book_tag)+"/book?start=0"
    try:
        res= requests.get(url=test_url,headers=has[random.randint(0,2)])
        if res.status_code != 200:
            raise "请求失败！"
    except Exception as e:
        print(e)
    bs = BeautifulSoup(res.text,'lxml')
    tag = bs.find("div",id="content").find_next("h1").string
    #获取书单列表div
    book_info=bs.find("div",class_="mod book-list").find_all("dd")
    for book in list(book_info):
        a_t =book.find("a",class_="title")
        title =a_t.string
        link=a_t.attrs['href']
        desc_list=book.find("div",class_="desc").string.strip()
        authors = desc_list.split('/')[:-3]
        if len(authors)>1:
            author = "作家:"+",".join(authors)
        else:
            author ="作家:"+desc_list.split('/')[0].strip()
        price = "价格："+desc_list.split('/')[-1].strip()
        product= "出版社："+desc_list.split('/')[-3].strip()
        print()
        rating= "评分："+book.find("span", class_="rating_nums").string
        book_list.append([title,link,author,price,product,rating])
    return book_list,tag

def write_exl(list,tag):
    exl_head =['书名','详情链接','作家','价格','出版社','评分']
    if not os.path.exists("豆瓣书单.xlsx"):
        wb =Workbook("豆瓣书单.xlsx")
        wb.save("豆瓣书单.xlsx")
    wb1=openpyxl.load_workbook("豆瓣书单.xlsx")
    #创建表页
    sheet=wb1.active
    sheet.title=tag
    #写入表头
    for i in range(len(exl_head)):
        sheet.cell(row=1,column=i+1,value=exl_head[i])
    #写入书单数据,从第二行开始
    for i in range(2,len(list)+2):
        for j in range(1,len(list[0])+1):
            print(list[i-2][j-1])
            sheet.cell(row=i,column=j,value=list[i-2][j-1])
    wb1.save("豆瓣书单.xlsx")





if __name__ == "__main__":
    book_list,tag =book_spider("小说")
    print(book_list)
    print(tag)
    write_exl(book_list,tag)