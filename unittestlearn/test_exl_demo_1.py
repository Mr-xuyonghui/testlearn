from openpyxl import *
wb =load_workbook(r'f:\useracount.xlsx')
#获取当前活动的表页
sheet = wb.active
print("exl所有的表页",wb.sheetnames)
#打印第一行的单元格对象
print(sheet[1])
#打印a列的单元格
print(sheet['A'])

#切换表页
#使用get_sheet_by_name()
#sheet =wb.get_sheet_by_name('Sheet1')
#使用表页名称切换
#sheet =wb['Sheet1']
#使用表页索引切换，从0开始
#sheet =wb[0]
#print("当前活动表页为：",sheet)

#获取单元格的值
#通过指定单元格使用.value取得对应的值
print("a1单元格的值：",sheet['A1'].value)
#通过cell（）方法定位单元格使用.value取得对应的值
print("a1单元格的值：",sheet.cell(1,1).value)

#获得某行或者某列的值
rows =sheet[1]
print("打印第一行所有值:",end=" ")
for row in  rows:
    print(row.value,end=" ")

#修改单元格的值
print("修改前a2单元格的值：",sheet['A2'].value)
sheet['A2']="xujiahui"
print("修改后a2单元格的值：",sheet['A2'].value)
#cell(行，列)
sheet.cell(2,1,"xurihui")
print("二次修改之后a2的值",sheet['A2'].value)

#添加一整行数据
rows =list(sheet.rows)[1:]
data =[]
for row in rows:
    row_data= []
    for cell in  row:
        row_data.append(cell.value)
    data.append(row_data)
print("整行添加前除表头外所有数据：",data)
sheet.append(["xujiahui",569,"男"])
rows =list(sheet.rows)[1:]
data =[]
for row in rows:
    row_data= []
    for cell in  row:
        row_data.append(cell.value)
    data.append(row_data)
print("整行添加后除表头外所有数据：",data)


