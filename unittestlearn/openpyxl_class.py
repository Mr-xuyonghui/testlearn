from  openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
class ExlHander():
    def __init__(self,filename):
        self.filename = filename
    def open_sheet(self,name) -> Worksheet :
        wb = load_workbook(self.filename)
        #通过类和表页名获取表页
        sheet = wb[name]
        #通过方法获取表页
        #sheet = wb.get_sheet_by_name(name)
        return sheet

    def header(self,sheet_name):
        '''获取表单的表头'''
        sheet = self.open_sheet(sheet_name)
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        return headers
    def read_list(self,sheet_name):
        sheet = self.open_sheet(sheet_name)
        #去除表头
        rows =list(sheet.rows)[1:]
        data =[]
        for row in rows:
            data_row =[]
            for cell in row:
                data_row.append(cell.value)
            data.append(data_row)
        return data
    def read_dict(self,sheet_name):
        sheet =self.open_sheet(sheet_name)
        rows =list(sheet.rows)[1:]
        data = []
        for row in rows:
            data_row = []
            for cell in row:
                data_row.append(cell.value)
                data_dict = dict(zip(self.header(sheet_name),data_row))
            data.append(data_dict)
        return data
    #通过实例修改
    def changa_cell(self,sheet_name,row,col,value):
        sheet =self.open_sheet(sheet_name)
        sheet.cell(row,col,value)
        #sheet.parent为sheet所属的工作布对象
        sheet.parent.save(self.filename)
    @staticmethod
    def sta_change_cell(file,sheet_name,row,col,value):
        wb =load_workbook(file)
        sheet =wb[sheet_name]
        sheet.cell(row,col,value)
        wb.save(file)
        wb.close()
if __name__ == '__main__':
    wb=ExlHander(r'f:\useracount.xlsx')
    # print(wb)
    # headers = wb.header('Sheet1')
    # print(headers)
    # print(wb.read_list('Sheet1'))
    # print(wb.read_dict('Sheet1'))
    #进行保存操作时，需要确保文件处于关闭状态
    wb.changa_cell('Sheet1',2,1,"xu111")
    #wb.sta_change_cell(r'f:\useracount.xlsx','Sheet1',2,1,"xu")

    # print(wb.read_list('Sheet1'))
